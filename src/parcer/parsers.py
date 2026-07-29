from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from random import sample, randint
import os
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import create_engine, insert
from .database import ProductOrder, Product, Order, Base
from datetime import datetime
import logging


class Parser:
    """ Главный класс для работы
    """
    def __init__(self, site_url, login, password, download_dir,
                 items_count, first_name, last_name, postal_code, 
                 db_name, log_filename, xlsx_filename):
        """ Инициализация парсера

        :param site_url: Ссылка на сайт
        :type site_url: str
        :param login: Логин пользователя
        :type login: str
        :param password: Пароль пользователя
        :type password: str
        :param download_dir: Папка, в которую будут загружаться файлы
        :type download_dir: str
        :param items_count: Количество элементов, которые будут добавлены в корзину. 
                            Если число больше имеющихся товаров, то будут добавлены все товары
        :type items_count: int
        :param first_name: Имя заказчкиа
        :type first_name: str
        :param last_name: Фамилия заказчика
        :type last_name: str
        :param postal_code: Почтовый индекс/ZIP-код
        :type postal_code: str
        :param db_name: Имя базы данных в формате filename.db
        :type db_name: str
        :param log_filename: Имя файла логов в формате filename.log
        :type log_filename: str
        :param xlsx_filename: Имя файла отчёта в формате filename.xlsx
        :type xlsx_filename: str
        """    

        self.site_url = site_url
        self.login = login
        self.password = password
        self.download_dir = download_dir
        self.items_count = items_count
        self.first_name = first_name
        self.last_name = last_name
        self.postal_code = postal_code
        self.db_name = db_name
        self.log_filename = log_filename
        self.xlsx_filename = xlsx_filename
        self.driver = self.create_driver()
        self.engine = self.create_database()
        self.logger = self.create_logger()
        self.logger.info('Successful parser initialization')

    def create_driver(self):
        """ Функция для создания главного драйвера.

        :return: Драйвер программы
        :rtype: selenium.webdriver.Chrome
        """ 

        options = Options()
        prefs = {
            "profile.password_manager_leak_detection": False,
            "download.default_directory": os.path.abspath(os.getcwd()) + self.download_dir,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)
        options.add_argument("-headless=new")
        options.add_argument("--no-sandbox")            
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")

        driver = webdriver.Chrome(options=options)
        return driver

    def create_database(self):
        """Функция для создания базы данных

        :return: Машина для работы с базой данных
        :rtype: sqlalchemy.engine.Engine
        """

        engine = create_engine(f'sqlite:///Databases/{self.db_name}')
        Base.metadata.create_all(engine)
        return engine

    def create_logger(self):
        """Функция для создания логгера программы

        :return: Логгер программы
        :rtype: logging.Logger
        """  

        logger = logging.getLogger('ProgramLogger')
        logger.setLevel(logging.INFO)

        file_handler = logging.FileHandler(f'logs/{self.log_filename}', 'a')
        file_handler.setLevel(logging.INFO)

        log_format = logging.Formatter('%(asctime)s %(name)s %(levelname)s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        file_handler.setFormatter(log_format)

        logger.addHandler(file_handler)
        return logger

    def authorization(self):
        """ Функция для авторизации """

        self.driver.find_element(By.ID, 'user-name').send_keys(self.login)
        self.driver.find_element(By.ID, 'password').send_keys(self.password)
        self.driver.find_element(By.ID, 'login-button').click()

    def create_cart(self):
        """ Формирование корзины. В корзину добавляются то количество случайных элементов,
        которые указал пользователь. Добавленные элементы сохраняются в базу данных
        """

        WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.ID, 'root')))

        items = self.driver.find_element(By.CLASS_NAME, 'inventory_list').find_elements(
            By.CLASS_NAME, 'inventory_item')
        if self.items_count > len(items):
            self.items_count = len(items)
        buy_items_indecies = sorted(
            sample(range(len(items)), self.items_count))
        k = 0

        self.current_items = []
        with Session(bind=self.engine) as session:
            for index, item in enumerate(items):
                if index == buy_items_indecies[k]:
                    title = item.find_element(
                        By.CLASS_NAME, 'inventory_item_name').text
                    description = item.find_element(
                        By.CLASS_NAME, 'inventory_item_desc').text
                    price = float(item.find_element(
                        By.CLASS_NAME, 'inventory_item_price').text[1:])
                    item.find_element(By.CLASS_NAME, 'btn').click()
                    res = session.query(Product).filter_by(
                        title=title, description=description, price=price)
                    if res.count() == 0:
                        product = Product(
                            title=title, description=description, price=price)
                        session.add(product)
                        session.commit()
                        self.current_items.append(product.id)
                    else:
                        self.current_items.append(res.first().id)
                    k += 1
                    if k == self.items_count:
                        break

    def delete_item_from_database(self):
        """ Удаление случайного элемента из корзины

        :return: Наименование удалённого элемента
        :rtype: str
        """

        delete_item_index = randint(0, self.items_count - 1)
        delete_item_id_on_database = self.current_items.pop(delete_item_index)
        with Session(bind=self.engine) as session:
            delete_item_title = session.query(Product).filter_by(
                id=delete_item_id_on_database).first().title
        return delete_item_title

    def go_to_cart(self):
        """ Переход в корзину """ 

        self.driver.find_element(By.CLASS_NAME, 'shopping_cart_link').click()

    def delete_item_from_cart(self, delete_item_title):
        """ Функция для удаления случайно выбранного элемента

        :param delete_item_title: Наименование удалённого элемента
        :type delete_item_title: str
        """

        WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.ID, 'root')))
        cart_items = self.driver.find_element(
            By.CLASS_NAME, 'cart_list').find_elements(By.CLASS_NAME, 'cart_item')
        for item in cart_items:
            if item.find_element(By.CLASS_NAME, 'inventory_item_name').text == delete_item_title:
                item.find_element(By.CLASS_NAME, 'btn').click()
                break

    def checkout_cart(self):
        """ Подтверждение товаров в корзине """

        self.driver.find_element(By.ID, 'checkout').click()

    def fill_user_data(self):
        """ Заполнение данных пользователя """

        WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.ID, 'root')))
        self.driver.find_element(
            By.ID, 'first-name').send_keys(self.first_name)
        self.driver.find_element(By.ID, 'last-name').send_keys(self.last_name)
        self.driver.find_element(
            By.ID, 'postal-code').send_keys(self.postal_code)
        self.driver.find_element(By.ID, 'continue').click()

    def get_payment_information(self):
        """ Получение информации о доставке и сохранение её в базу данных """

        WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.ID, 'root')))
        info_item_block = self.driver.find_element(
            By.CLASS_NAME, 'summary_info')
        info_items = info_item_block.find_elements(
            By.CLASS_NAME, 'summary_value_label')
        self.payment_information = info_items[0].text
        self.shipping_information = info_items[1].text
        self.tax = float(self.driver.find_element(
            By.CLASS_NAME, 'summary_tax_label').text.split()[1][1:])
        self.total = float(self.driver.find_element(
            By.CLASS_NAME, 'summary_total_label').text.split()[1][1:])
        info_item_block.find_element(By.ID, 'finish').click()
        with Session(bind=self.engine) as session:
            order = Order(first_name=self.first_name, last_name=self.last_name,
                          postal_code=self.postal_code, payment_information=self.payment_information,
                          shipping_information=self.shipping_information, tax=self.tax, price_on_cite=self.total)
            session.add(order)
            session.commit()
            insert_data = [
                {"orderId": order.id, "productId": product_id}
                for product_id in self.current_items
            ]
            self.current_order_id = order.id
            session.execute(insert(ProductOrder), insert_data)
            session.commit()

    def create_pdf(self):
        """ Получение отчёта с сервера сайта """

        WebDriverWait(self.driver, 30).until(
            EC.element_to_be_clickable((By.ID, 'generate-pdf-order')))
        self.driver.find_element(By.ID, 'generate-pdf-order').click()
        WebDriverWait(self.driver, 30).until(EC.text_to_be_present_in_element(
            (By.ID, 'generate-pdf-order'), 'Generate PDF order'))
        self.driver.quit()

    def set_header(self, cells, name, value, border):
        """ Создание заголовков в отчёте

        :param cells: Ячейки, куда будут записаны данные
        :type cells: List[str, str]
        :param name: Заголовок ячейки
        :type name: str
        :param value: Значение ячейки
        :type value: str
        :param border: Тип рамки для ячеек
        :type border: openpyxl.styles.Border
        """   

        self.work_sheet[cells[0]] = name
        self.work_sheet[cells[0]].border = border
        self.work_sheet[cells[1]] = value
        self.work_sheet[cells[1]].border = border
        self.work_sheet[cells[1]].alignment = Alignment(
            wrap_text=True, horizontal='justify')

    def set_merge_rows_value(self, cells, name, value, column, start_row, end_row):
        """ Объединение ячеек и установка значение в ней

        :param cells: Ячейки, в которых необходимо установить значения
        :type cells: List[str, str]
        :param name: Заголовок ячейки
        :type name: str
        :param value: Значение ячейки
        :type value: str/int
        :param column: Объединяемая колонка
        :type column: int
        :param start_row: Начальная строка для объединения
        :type start_row: int
        :param end_row: Конечная строка для объединения
        :type end_row: int
        """

        self.work_sheet.merge_cells(
            start_row=start_row, start_column=column, end_row=end_row, end_column=column)
        self.work_sheet[cells[0]] = name
        self.work_sheet[cells[1]] = value

    def set_column_width(self, values):
        """ Установление ширины колонок

        :param values: Словарь со значениями ширины столбцов. Например, {'A': 10, 'B': 20}
        :type values: dict
        """

        for key, value in values.items():
            self.work_sheet.column_dimensions[key].width = value

    def create_equal_not_equal_color_rule(self, cell, compare_cells):
        """ Создание правила для сравнения ячеек.
            Если ячейки равны, то будет окрашено в зеленый цвет, если нет, то в красный

        :param cell: Ячейка, в которую будет записана формула
        :type cell: str
        :param compare_cells: Сравниваемые ячейки
        :type compare_cells: List[str, str]
        """
        
        red_fill = PatternFill(start_color='FF0000', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', fill_type='solid')

        rule_equal = FormulaRule(
            formula=[f'{compare_cells[0]}={compare_cells[1]}'], fill=green_fill)
        rule_not_equal = FormulaRule(
            formula=[f'{compare_cells[0]}<>{compare_cells[1]}'], fill=red_fill)

        self.work_sheet.conditional_formatting.add(cell, rule_equal)
        self.work_sheet.conditional_formatting.add(cell, rule_not_equal)

    def create_table_borders(self, head_range, head_border, table_range, table_border):
        """ Создание границы для главной таблицы

        :param head_range: Границы заголовка таблицы. Например, 'A4:I4'
        :type head_range: str
        :param head_border: Тип границы для заголовка таблицы
        :type head_border: openpyxl.styles.Border
        :param table_range: Границы данных таблицы таблицы. Например, 'A5:I6'
        :type table_range: str
        :param table_border: Тип границы для данных таблицы
        :type table_border: openpyxl.styles.Border
        """

        for row in self.work_sheet[table_range]:
            for cell in row:
                cell.border = table_border

        for row in self.work_sheet[head_range]:
            for cell in row:
                cell.border = head_border

    def set_small_text_alignment(self, range):
        """ Создание выравнивания для небольшого текста

        :param range: Ячейки, в которых необходимо выставить выравнивание. Например, 'A4:I6'
        :type range: str
        """

        for row in self.work_sheet[range]:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

    def set_large_text_alignment(self, range):
        """ Создание выравнивания для объёмного текста

        :param range: Ячейки, в которых необходимо выставить выравнивание. Например, 'C5:C6'
        :type range: str
        """

        for col in self.work_sheet[range]:
            for cell in col:
                cell.alignment = Alignment(
                    wrap_text=True, horizontal='justify')

    def set_main_data(self, header, data, range):
        """ Заполнение данных таблицы из базы данных

        :param header: Заголовок таблицы
        :type header: tuple
        :param data: Главные данные
        :type data: List[tuple]
        :param range: Ячейки, в которые необходимо вставить данные. Например, 'A4:D6'
        :type range: str
        """

        data = [header] + data
        data = [item for sublist in data for item in sublist]
        k = 0
        for row in self.work_sheet[range]:
            for cell in row:
                cell.value = data[k]
                k += 1

    def create_xlsx_file(self):
        """ Создание XLSX отчёта """

        self.work_book = Workbook()
        self.work_sheet = self.work_book.active
        self.work_sheet.title = 'Report'

        width_values = {
            'B': 30,
            'C': 30,
            'D': 30,
            'E': 30,
            'F': 15,
            'G': 15,
            'H': 15,
            'I': 15
        }

        thin_side = Side(border_style="thin", color="000000")
        table_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )
        head_side = Side(border_style="medium", color="000000")
        head_border = Border(
            left=head_side, right=head_side, top=head_side, bottom=head_side
        )

        with Session(bind=self.engine) as session:
            order_info = session.query(Order).filter_by(
                id=self.current_order_id).first()
            product_count_in_order = session.query(
                ProductOrder).filter_by(orderId=order_info.id).count()
            products_in_order = order_info.products
            products_data = [(index + 1, item.title, item.description, item.price)
                             for index, item in enumerate(products_in_order)]
            session.close()

        self.set_header(['B1', 'B2'], 'Customer',
                        f'{order_info.first_name} {order_info.last_name}', table_border)
        self.set_header(['C1', 'C2'], 'ZIP/Post code',
                        order_info.postal_code, table_border)
        self.set_header(['D1', 'D2'], 'Payment Information',
                        order_info.payment_information, table_border)
        self.set_header(['E1', 'E2'], 'Shipping Information',
                        order_info.shipping_information, table_border)

        self.set_column_width(width_values)

        self.set_main_data(('№', 'Product', 'Description', 'Price'),
                           products_data, f'A4:D{4 + product_count_in_order}')

        self.set_merge_rows_value(
            ['E4', 'E5'], 'Item total', f'=SUM(D5:D{4 + product_count_in_order})', 5, 5, 4 + product_count_in_order)
        self.set_merge_rows_value(
            ['F4', 'F5'], 'Tax', order_info.tax, 6, 5, 4 + product_count_in_order)
        self.set_merge_rows_value(
            ['G4', 'G5'], 'Total', '=E5 + F5', 7, 5, 4 + product_count_in_order)
        self.set_merge_rows_value(['H4', 'H5'], 'Total on cite',
                                  order_info.price_on_cite, 8, 5, 4 + product_count_in_order)

        self.set_merge_rows_value(
            ['I4', 'I5'], 'Equal totals', '', 9, 5, 4 + product_count_in_order)
        self.create_equal_not_equal_color_rule('I5', ['$H$5', '$G$5'])

        self.create_table_borders(
            'A4:I4', head_border, f'A5:I{4 + product_count_in_order}', table_border)

        self.set_small_text_alignment(f'A4:I{4 + product_count_in_order}')
        self.set_large_text_alignment(f'C5:C{4 + product_count_in_order}')

        self.work_book.save(f'{os.path.abspath(os.getcwd())}{self.download_dir}/{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}_{self.xlsx_filename}')

    def start(self):
        """ Главная функция """        
        self.driver.get(self.site_url)

        #!--------task1--------!
        self.authorization()
        self.logger.info('Successful authorization')

        #!--------task2-3--------!
        self.create_cart()
        self.logger.info('Items added to the cart')

        #!--------task4--------!
        self.delete_item_title = self.delete_item_from_database()

        #!--------task5--------!
        self.go_to_cart()

        #!--------task6--------!
        self.delete_item_from_cart(self.delete_item_title)
        self.logger.info('A randomly selected item has been removed from the cart')

        #!--------task7--------!
        self.checkout_cart()
        self.logger.info('The shopping cart has been checked')

        #!--------task8--------!
        self.fill_user_data()
        self.logger.info('User data has been entered')

        #!--------task9--------!
        self.get_payment_information()
        self.logger.info('Delivery information has been filled in')

        #!--------task10--------!
        self.create_pdf()
        self.logger.info('PDF file uploaded')

        #!--------task11--------!
        self.create_xlsx_file()
        self.logger.info('XLSX report created')
