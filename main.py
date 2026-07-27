from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from random import sample, randint
from time import sleep
import os
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo




site_url = 'https://www.saucedemo.com/'
login = 'standard_user'
password = 'secret_sauce'
download_dir = os.path.abspath(os.getcwd()) + '/Downloads'

options = Options()
prefs = {
        "profile.password_manager_leak_detection": False,
        "download.default_directory": download_dir,
        "safebrowsing.enabled": True,
}
options.add_experimental_option("prefs", prefs)
#options.add_argument('--headless=new')

driver = webdriver.Chrome(options=options)
driver.get(site_url)

#!--------task1--------!
driver.find_element(By.ID, 'user-name').send_keys(login)
driver.find_element(By.ID, 'password').send_keys(password)
driver.find_element(By.ID, 'login-button').click()

#!--------task2-3--------!
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))

items_count = 4
items = driver.find_element(By.CLASS_NAME, 'inventory_list').find_elements(By.CLASS_NAME, 'inventory_item')
buy_items_indecies = sorted(sample(range(len(items)), items_count))
k = 0

database = []
for index, item in enumerate(items):
    if index == buy_items_indecies[k]:
        title = item.find_element(By.CLASS_NAME, 'inventory_item_name').text 
        description = item.find_element(By.CLASS_NAME, 'inventory_item_desc').text 
        price = float(item.find_element(By.CLASS_NAME, 'inventory_item_price').text[1:])
        item.find_element(By.CLASS_NAME, 'btn').click()
        database.append((title, description, price))
        k += 1
        if k == items_count:
            break

#!--------task4--------!
delete_item_index = randint(0, items_count - 1)
delete_item = database.pop(delete_item_index)

#!--------task5--------!
driver.find_element(By.CLASS_NAME, 'shopping_cart_link').click()

#!--------task6--------!
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))
cart_items = driver.find_element(By.CLASS_NAME, 'cart_list').find_elements(By.CLASS_NAME, 'cart_item')
for item in cart_items:
    if item.find_element(By.CLASS_NAME, 'inventory_item_name').text == delete_item[0]:
        item.find_element(By.CLASS_NAME, 'btn').click()
        break

#!--------task7--------!
driver.find_element(By.ID, 'checkout').click()

#!--------task8--------!
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))
first_name = 'Evgeny'
last_name = 'Kurshev'
postal_code = '187550'
driver.find_element(By.ID, 'first-name').send_keys(first_name)
driver.find_element(By.ID, 'last-name').send_keys(last_name)
driver.find_element(By.ID, 'postal-code').send_keys(postal_code)
driver.find_element(By.ID, 'continue').click()

#!--------task9--------!
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))
info_item_block = driver.find_element(By.CLASS_NAME, 'summary_info')
info_items = info_item_block.find_elements(By.CLASS_NAME, 'summary_value_label')
payment_information = info_items[0].text
shipping_information = info_items[1].text
tax = float(driver.find_element(By.CLASS_NAME, 'summary_tax_label').text.split()[1][1:])
total = float(driver.find_element(By.CLASS_NAME, 'summary_total_label').text.split()[1][1:])
info_item_block.find_element(By.ID, 'finish').click()

#!--------task10--------!
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))
driver.find_element(By.ID, 'generate-pdf-order').click()

#!--------task11--------!
work_book = Workbook()
work_sheet = work_book.active
work_sheet.title = 'Report'
work_sheet['B1'] = 'Customer'
work_sheet.column_dimensions['B'].width = 30
work_sheet['B2'] = f'{first_name} {last_name}'
work_sheet['C1'] = 'ZIP/Post code'
work_sheet['C2'] = postal_code

data = [('№', 'Product', 'Description', 'Price')] + [(index, *item) for index, item in enumerate(database)]
data = [item for sublist in data for item in sublist]
k = 0
for row in work_sheet.iter_rows(min_row=4, max_row=3 + items_count, min_col=1, max_col=4):
    for cell in row:
        cell.value = data[k]
        k += 1

work_sheet.column_dimensions['C'].width = 30

for column in ['D', 'E', 'F', 'G', 'H', 'I']:
    work_sheet.column_dimensions[column].width = 15

work_sheet.merge_cells(start_row=5, start_column=5, end_row=3 + items_count, end_column=5)
work_sheet['E4'] = 'Item total'
work_sheet['E5'] = f'=SUM(D5:D{3 + items_count})'

work_sheet.merge_cells(start_row=5, start_column=6, end_row=3 + items_count, end_column=6)
work_sheet['F4'] = 'Tax'
work_sheet['F5'] = tax

work_sheet.merge_cells(start_row=5, start_column=7, end_row=3 + items_count, end_column=7)
work_sheet['G4'] = 'Total'
work_sheet['G5'] = '=E5 + F5'

work_sheet.merge_cells(start_row=5, start_column=8, end_row=3 + items_count, end_column=8)
work_sheet['H4'] = 'Total on cite'
work_sheet['H5'] = total

work_sheet.merge_cells(start_row=5, start_column=9, end_row=3 + items_count, end_column=9)
work_sheet['I4'] = 'Equal totals'

red_fill = PatternFill(start_color='FF0000', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', fill_type='solid')

rule_equal = FormulaRule(formula=['$H$5=$G$5'], fill=green_fill)
rule_not_equal = FormulaRule(formula=['$H$5<>$G$5'], fill=red_fill)

work_sheet.conditional_formatting.add('I5', rule_equal)
work_sheet.conditional_formatting.add('I5', rule_not_equal)

thin_side = Side(border_style="thin", color="000000")
table_border = Border(
    left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
)
head_side = Side(border_style="medium", color="000000")
head_border = Border(
    left=head_side, right=head_side, top=head_side, bottom=head_side
)

for row in work_sheet["B1:C2"]:
    for cell in row:
        cell.border = table_border

for row in work_sheet[f'A5:I{3 + items_count}']:
    for cell in row:
        cell.border = table_border

for row in work_sheet["A4:I4"]:
    for cell in row:
        cell.border = head_border

for row in work_sheet[f'A4:I{3 + items_count}']:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col in work_sheet.iter_rows(min_col=3, max_col=3, min_row=5, max_row=4 + items_count):
    for cell in col:
        cell.alignment = Alignment(wrap_text=True, horizontal='justify')


work_book.save('Report.xlsx')



sleep(30)

