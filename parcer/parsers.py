from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from random import sample, randint
import os
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment, Border, Side


class Parser:

    def __init__(self, site_url, login, password, download_dir, 
                    items_count, first_name, last_name, postal_code):
        self.site_url = site_url
        self.login = login
        self.password = password
        self.download_dir = download_dir
        self.items_count = items_count
        self.first_name = first_name
        self.last_name = last_name
        self.postal_code = postal_code
        self.driver = self.create_driver()
        

    def create_driver(self):
        options = Options()
        prefs = {
            "profile.password_manager_leak_detection": False,
            "download.default_directory": os.path.abspath(os.getcwd()) + self.download_dir,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)
        #options.add_argument('--headless=new')

        driver = webdriver.Chrome(options=options)
        return driver

    def authorization(self, ):
        self.driver.find_element(By.ID, 'user-name').send_keys(self.login)
        self.driver.find_element(By.ID, 'password').send_keys(self.password)
        self.driver.find_element(By.ID, 'login-button').click()

    def create_cart(self):
        WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))

        items = self.driver.find_element(By.CLASS_NAME, 'inventory_list').find_elements(By.CLASS_NAME, 'inventory_item')
        buy_items_indecies = sorted(sample(range(len(items)), self.items_count))
        k = 0

        database = []
        for index, item in enumerate(items):
            if index == buy_items_indecies[k]:
                title = item.find_element(By.CLASS_NAME, 'inventory_item_name').text 
                description = item.find_element(By.CLASS_NAME, 'inventory_item_desc').text 
                price = float(item.find_element(By.CLASS_NAME, 'inventory_item_price').text[1:])
                item.find_element(By.CLASS_NAME, 'btn').click()
                database.append((title, description, price))
                k += 1
                if k == self.items_count:
                    break

        return database

    def delete_item_from_database(self):
        delete_item_index = randint(0, self.items_count - 1)
        delete_item = self.database.pop(delete_item_index)
        return delete_item

    def go_to_cart(self):
        self.driver.find_element(By.CLASS_NAME, 'shopping_cart_link').click()

    def delete_item_from_cart(self, delete_item):
        WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))
        cart_items = self.driver.find_element(By.CLASS_NAME, 'cart_list').find_elements(By.CLASS_NAME, 'cart_item')
        for item in cart_items:
            if item.find_element(By.CLASS_NAME, 'inventory_item_name').text == delete_item[0]:
                item.find_element(By.CLASS_NAME, 'btn').click()
                break

    def checkout_cart(self):
        self.driver.find_element(By.ID, 'checkout').click()

    def fill_user_data(self):
        WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))
        self.driver.find_element(By.ID, 'first-name').send_keys(self.first_name)
        self.driver.find_element(By.ID, 'last-name').send_keys(self.last_name)
        self.driver.find_element(By.ID, 'postal-code').send_keys(self.postal_code)
        self.driver.find_element(By.ID, 'continue').click()

    def get_payment_information(self):
        WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, 'root')))
        info_item_block = self.driver.find_element(By.CLASS_NAME, 'summary_info')
        info_items = info_item_block.find_elements(By.CLASS_NAME, 'summary_value_label')
        payment_information = info_items[0].text
        shipping_information = info_items[1].text
        tax = float(self.driver.find_element(By.CLASS_NAME, 'summary_tax_label').text.split()[1][1:])
        total = float(self.driver.find_element(By.CLASS_NAME, 'summary_total_label').text.split()[1][1:])
        info_item_block.find_element(By.ID, 'finish').click()
        return payment_information, shipping_information, tax, total

    def create_pdf(self):
        WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.ID, 'generate-pdf-order')))
        self.driver.find_element(By.ID, 'generate-pdf-order').click()
        WebDriverWait(self.driver, 30).until(EC.text_to_be_present_in_element((By.ID, 'generate-pdf-order'), 'Generate PDF order'))
        self.driver.quit()

    def set_header(self, cells, name, value, border):
        self.work_sheet[cells[0]] = name
        self.work_sheet[cells[0]].border = border
        self.work_sheet[cells[1]] = value
        self.work_sheet[cells[1]].border = border
        self.work_sheet[cells[1]].alignment = Alignment(wrap_text=True, horizontal='justify')

    def set_merge_rows_value(self, cells, name, value, column, start_row, end_row):
        self.work_sheet.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)
        self.work_sheet[cells[0]] = name
        self.work_sheet[cells[1]] = value

    def set_column_width(self, values):
        for key, value in values.items():
            self.work_sheet.column_dimensions[key].width = value

    def create_equal_not_equal_color_rule(self, cell, compare_cells):
        red_fill = PatternFill(start_color='FF0000', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', fill_type='solid')

        rule_equal = FormulaRule(formula=[f'{compare_cells[0]}={compare_cells[1]}'], fill=green_fill)
        rule_not_equal = FormulaRule(formula=[f'{compare_cells[0]}<>{compare_cells[1]}'], fill=red_fill)

        self.work_sheet.conditional_formatting.add(cell, rule_equal)
        self.work_sheet.conditional_formatting.add(cell, rule_not_equal)

    def create_table_borders(self, head_range, head_border, table_range, table_border):
        for row in self.work_sheet[table_range]:
            for cell in row:
                cell.border = table_border

        for row in self.work_sheet[head_range]:
            for cell in row:
                cell.border = head_border

    def set_small_text_alignment(self, range):
        for row in self.work_sheet[range]:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def set_large_text_alignment(self, range):
        for col in self.work_sheet[range]:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, horizontal='justify')

    def set_main_data(self, header, data, range):
        data = [header] + [(index + 1, *item) for index, item in enumerate(data)]
        data = [item for sublist in data for item in sublist]
        k = 0
        for row in self.work_sheet[range]:
            for cell in row:
                cell.value = data[k]
                k += 1
     
    def create_xlsx_file(self):
        self.work_book = Workbook()
        self.work_sheet = self.work_book.active
        self.work_sheet.title = 'Report'

        width_values = {
            'B': 30,
            'C': 30,
            'D': 30,
            'E': 30,
            'F': 15,
            'G': 15,
            'H': 15,
            'I': 15
        }

        thin_side = Side(border_style="thin", color="000000")
        table_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )
        head_side = Side(border_style="medium", color="000000")
        head_border = Border(
            left=head_side, right=head_side, top=head_side, bottom=head_side
        )

        self.set_header(['B1', 'B2'], 'Customer', f'{self.first_name} {self.last_name}', table_border)
        self.set_header(['C1', 'C2'], 'ZIP/Post code', self.postal_code, table_border)
        self.set_header(['D1', 'D2'], 'Payment Information', self.payment_information, table_border)
        self.set_header(['E1', 'E2'], 'Shipping Information', self.shipping_information, table_border)

        self.set_column_width(width_values)

        self.set_main_data(('№', 'Product', 'Description', 'Price'), self.database, f'A4:D{3 + self.items_count}')

        self.set_merge_rows_value(['E4', 'E5'], 'Item total', f'=SUM(D5:D{3 + self.items_count})', 5, 5, 3 + self.items_count)
        self.set_merge_rows_value(['F4', 'F5'], 'Tax', self.tax, 6, 5, 3 + self.items_count)
        self.set_merge_rows_value(['G4', 'G5'], 'Total', '=E5 + F5', 7, 5, 3 + self.items_count)
        self.set_merge_rows_value(['H4', 'H5'], 'Total on cite', self.total, 8, 5, 3 + self.items_count)

        self.set_merge_rows_value(['I4', 'I5'], 'Equal totals', '', 9, 5, 3 + self.items_count)
        self.create_equal_not_equal_color_rule('I5', ['$H$5', '$G$5'])

        self.create_table_borders('A4:I4', head_border, f'A5:I{3 + self.items_count}', table_border)

        self.set_small_text_alignment(f'A4:I{3 + self.items_count}')
        self.set_large_text_alignment(f'C5:C{3 + self.items_count}')

        self.work_book.save('Report.xlsx')

    def start(self):
        self.driver.get(self.site_url)

        #!--------task1--------!
        self.authorization()

        #!--------task2-3--------!
        self.database = self.create_cart()

        #!--------task4--------!
        self.delete_item = self.delete_item_from_database()

        #!--------task5--------!
        self.go_to_cart()

        #!--------task6--------!
        self.delete_item_from_cart(self.delete_item)

        #!--------task7--------!
        self.checkout_cart()

        #!--------task8--------!
        self.fill_user_data()

        #!--------task9--------!
        self.payment_information, self.shipping_information, self.tax, self.total = self.get_payment_information()

        #!--------task10--------!
        self.create_pdf()

        #!--------task11--------!
        self.create_xlsx_file()
    